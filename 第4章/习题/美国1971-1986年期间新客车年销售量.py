# coding=utf-8
# Author: 史浩 浙江金融职业学院
# --------------------------------------------------------
import pandas as pd
import numpy as np
import statsmodels.api as sm

df = pd.read_excel(r'..\..\data\美国1971-1986年期间新客车年销售量.xlsx', sheet_name="Sheet1")
df = df.iloc[:, :7]
df = df.dropna()
# print(df)
#              0      1    2     3     4     5     6
df.columns = ['year', 'y', 'x1', 'x2', 'x3', 'x4', 'x5']
df.reset_index(drop=True, inplace=True)  # 把索引重新排一下
print(df)

# 定义变量
# 用.astype(float)将EXCEL文件中的字符型转浮点型
y = df.iloc[:, 1].astype(float).apply(np.log)
x1 = df.iloc[:, 2].astype(float).apply(np.log)
x2 = df.iloc[:, 3].astype(float).apply(np.log)
x3 = df.iloc[:, 4].astype(float).apply(np.log)
x4 = df.iloc[:, 5].astype(float).apply(np.log)
x5 = df.iloc[:, 6].astype(float).apply(np.log)

Z = np.column_stack((x1, x2, x3, x4, x5))
##########################################################################
# 进行多元线性回归，可以替换Y和X
#          OLS：  Y          = beta0 + beta1*X1 + beta2*X2
##########################################################################
model = sm.OLS(y, sm.add_constant(Z))  # 用add_constant加入常数项
print(model.fit().summary(xname=['const', 'lnx1', 'lnx2', 'lnx3', 'lnx4', 'lnx5']))  # 用自己的名称命名常数和各个解释变量

# 相关系数
print(np.corrcoef((x1, x2, x3, x4, x5)))

# 仅打印系数
res = model.fit()
print(res.params)

df_out = pd.concat((res.params, res.tvalues, res.pvalues, res.conf_int()), axis=1)
df_out.columns = ['beta', 't', 'p-value', 'CI_low', 'CI_high']
print(df_out)
# 输出到文件
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
wb = Workbook()
ws = wb.create_sheet("sheet1")
rows = dataframe_to_rows(df_out)

for r_idx, row in enumerate(rows, 1):
    for c_idx, value in enumerate(row, 1):
        ws.cell(row=r_idx, column=c_idx, value=value)
wb.save("output.xls")  # 输出到文件

